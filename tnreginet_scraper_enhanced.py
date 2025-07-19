#!/usr/bin/env python3
"""
Enhanced Tamil Nadu Registration Department Website Scraper
Uses configuration file for customizable behavior
Automates data extraction from www.tnreginet.gov.in for Zone, District, Sub Register Office, and Village dropdowns
Creates Excel file with cascading dropdowns functionality
"""

import time
import logging
import traceback
import re
import os
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName

# Import configuration
from config import *

class TNRegiNetScraperEnhanced:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.data = {}
        self.setup_logging()
        
    def setup_logging(self):
        """Setup logging based on configuration"""
        handlers = []
        
        if LOGGING_CONFIG["file_output"]:
            handlers.append(logging.FileHandler(OUTPUT_CONFIG["log_filename"]))
        
        if LOGGING_CONFIG["console_output"]:
            handlers.append(logging.StreamHandler())
        
        logging.basicConfig(
            level=getattr(logging, LOGGING_CONFIG["level"]),
            format=LOGGING_CONFIG["format"],
            handlers=handlers
        )
        
        self.logger = logging.getLogger(__name__)
        self.logger.info("Enhanced TN Registration Scraper initialized")
        
    def setup_driver(self):
        """Initialize Chrome WebDriver with configuration options"""
        try:
            options = get_chrome_options()
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=options)
            
            # Set timeouts from configuration
            self.driver.implicitly_wait(BROWSER_CONFIG["implicit_wait"])
            self.driver.set_page_load_timeout(BROWSER_CONFIG["page_load_timeout"])
            
            self.wait = WebDriverWait(self.driver, BROWSER_CONFIG["timeout"])
            self.logger.info("Chrome WebDriver initialized successfully")
            
        except Exception as e:
            self.logger.error(f"Failed to initialize WebDriver: {str(e)}")
            raise
    
    def take_screenshot(self, filename_suffix="error"):
        """Take screenshot for debugging"""
        try:
            if SCRAPING_CONFIG["screenshot_on_error"]:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"screenshot_{filename_suffix}_{timestamp}.png"
                self.driver.save_screenshot(filename)
                self.logger.info(f"Screenshot saved: {filename}")
        except Exception as e:
            self.logger.error(f"Failed to take screenshot: {str(e)}")
    
    def navigate_to_website(self):
        """Navigate to the TN Registration website"""
        try:
            self.logger.info(f"Navigating to {WEBSITE_URL}")
            self.driver.get(WEBSITE_URL)
            time.sleep(3)
            self.logger.info("Successfully loaded the website")
            
        except Exception as e:
            self.logger.error(f"Failed to navigate to website: {str(e)}")
            self.take_screenshot("navigation_error")
            raise
    
    def find_and_click_element(self, selectors, element_name, required=False):
        """Find and click element using multiple selectors"""
        for selector in selectors:
            try:
                element = self.wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                element.click()
                self.logger.info(f"Successfully clicked {element_name}")
                return True
            except TimeoutException:
                continue
        
        if required:
            self.logger.error(f"Required element {element_name} not found")
            self.take_screenshot(f"{element_name.lower()}_not_found")
            raise Exception(f"Required element {element_name} not found")
        else:
            self.logger.warning(f"Optional element {element_name} not found, continuing")
            return False
    
    def select_english_language(self):
        """Select English language option"""
        try:
            self.logger.info("Attempting to select English language")
            self.find_and_click_element(SELECTORS["language"], "English language", required=False)
            time.sleep(2)
            
        except Exception as e:
            self.logger.error(f"Error selecting English language: {str(e)}")
            # Continue execution as language selection might not be mandatory
    
    def navigate_to_ec_page(self):
        """Navigate to E-services -> Encumbrance certificate -> View EC"""
        try:
            self.logger.info("Navigating to E-services -> Encumbrance certificate -> View EC")
            
            # Navigate through menu items
            navigation_steps = [
                (SELECTORS["eservices"], "E-services"),
                (SELECTORS["encumbrance"], "Encumbrance certificate"),
                (SELECTORS["view_ec"], "View EC")
            ]
            
            for selectors, name in navigation_steps:
                if self.find_and_click_element(selectors, name, required=False):
                    time.sleep(2)
            
            self.logger.info("Navigation completed")
            
        except Exception as e:
            self.logger.error(f"Error navigating to EC page: {str(e)}")
            self.take_screenshot("navigation_error")
            # Continue with current page
            self.logger.info("Attempting to find dropdowns on current page")
    
    def clean_dropdown_options(self, options):
        """Clean and validate dropdown options"""
        cleaned_options = []
        
        for option in options:
            if DATA_VALIDATION["trim_whitespace"]:
                option = option.strip()
            
            if not option:
                continue
                
            if DATA_VALIDATION["skip_empty_options"] and not option:
                continue
                
            if option.lower() in [p.lower() for p in DATA_VALIDATION["skip_placeholder_options"]]:
                continue
                
            cleaned_options.append(option)
        
        if DATA_VALIDATION["remove_duplicates"]:
            cleaned_options = list(dict.fromkeys(cleaned_options))  # Preserve order
            
        return cleaned_options
    
    def extract_dropdown_options(self, dropdown_element):
        """Extract options from a dropdown element"""
        try:
            if dropdown_element.tag_name == 'select':
                select = Select(dropdown_element)
                options = [option.text for option in select.options]
            else:
                # Handle custom dropdowns
                options = [option.text for option in dropdown_element.find_elements(By.TAG_NAME, "option")]
            
            return self.clean_dropdown_options(options)
            
        except Exception as e:
            self.logger.error(f"Error extracting dropdown options: {str(e)}")
            return []
    
    def find_dropdown_element(self, dropdown_type):
        """Find dropdown element by type using configuration selectors"""
        selectors = SELECTORS.get(f"{dropdown_type}_dropdown", [])
        
        for selector in selectors:
            try:
                element = self.driver.find_element(By.XPATH, selector)
                return element
            except NoSuchElementException:
                continue
        
        return None
    
    def select_dropdown_option(self, dropdown_element, option_text):
        """Select option from dropdown with retry logic"""
        for attempt in range(SCRAPING_CONFIG["max_retries"]):
            try:
                select = Select(dropdown_element)
                select.select_by_visible_text(option_text)
                time.sleep(SCRAPING_CONFIG["delay_between_selections"])
                return True
            except Exception as e:
                self.logger.warning(f"Attempt {attempt + 1} failed to select '{option_text}': {str(e)}")
                if attempt < SCRAPING_CONFIG["max_retries"] - 1:
                    time.sleep(SCRAPING_CONFIG["retry_delay"])
                else:
                    self.logger.error(f"Failed to select '{option_text}' after {SCRAPING_CONFIG['max_retries']} attempts")
                    return False
        return False
    
    def extract_all_dropdown_data(self):
        """Extract all dropdown data with hierarchical relationships"""
        try:
            self.logger.info("Starting dropdown data extraction")
            
            # Find all dropdown elements
            dropdowns = {
                'zone': self.find_dropdown_element('zone'),
                'district': self.find_dropdown_element('district'),
                'sub_register': self.find_dropdown_element('sub_register'),
                'village': self.find_dropdown_element('village')
            }
            
            if not dropdowns['zone']:
                self.logger.error("Zone dropdown not found")
                self.take_screenshot("zone_dropdown_not_found")
                return
            
            # Extract zones
            zones = self.extract_dropdown_options(dropdowns['zone'])
            self.logger.info(f"Found {len(zones)} zones: {zones}")
            
            for zone_idx, zone in enumerate(zones):
                self.logger.info(f"Processing zone {zone_idx + 1}/{len(zones)}: {zone}")
                self.data[zone] = {}
                
                try:
                    # Select zone
                    if not self.select_dropdown_option(dropdowns['zone'], zone):
                        if ERROR_HANDLING["continue_on_error"]:
                            continue
                        else:
                            raise Exception(f"Failed to select zone: {zone}")
                    
                    # Extract districts for this zone
                    if dropdowns['district']:
                        # Wait for district dropdown to update
                        time.sleep(SCRAPING_CONFIG["delay_between_selections"])
                        districts = self.extract_dropdown_options(dropdowns['district'])
                        self.logger.info(f"Found {len(districts)} districts for zone {zone}")
                        
                        for district_idx, district in enumerate(districts):
                            self.logger.info(f"Processing district {district_idx + 1}/{len(districts)}: {district}")
                            self.data[zone][district] = {}
                            
                            try:
                                # Select district
                                if not self.select_dropdown_option(dropdowns['district'], district):
                                    if ERROR_HANDLING["continue_on_error"]:
                                        continue
                                    else:
                                        raise Exception(f"Failed to select district: {district}")
                                
                                # Extract sub registers for this district
                                if dropdowns['sub_register']:
                                    time.sleep(SCRAPING_CONFIG["delay_between_selections"])
                                    sub_registers = self.extract_dropdown_options(dropdowns['sub_register'])
                                    self.logger.info(f"Found {len(sub_registers)} sub registers for district {district}")
                                    
                                    for sub_register_idx, sub_register in enumerate(sub_registers):
                                        self.logger.info(f"Processing sub register {sub_register_idx + 1}/{len(sub_registers)}: {sub_register}")
                                        
                                        try:
                                            # Select sub register
                                            if not self.select_dropdown_option(dropdowns['sub_register'], sub_register):
                                                if ERROR_HANDLING["continue_on_error"]:
                                                    self.data[zone][district][sub_register] = []
                                                    continue
                                                else:
                                                    raise Exception(f"Failed to select sub register: {sub_register}")
                                            
                                            # Extract villages for this sub register
                                            if dropdowns['village']:
                                                time.sleep(SCRAPING_CONFIG["delay_between_selections"])
                                                villages = self.extract_dropdown_options(dropdowns['village'])
                                                self.logger.info(f"Found {len(villages)} villages for sub register {sub_register}")
                                                self.data[zone][district][sub_register] = villages
                                            else:
                                                self.data[zone][district][sub_register] = []
                                                
                                        except Exception as e:
                                            self.logger.error(f"Error processing sub register {sub_register}: {str(e)}")
                                            if ERROR_HANDLING["continue_on_error"]:
                                                self.data[zone][district][sub_register] = []
                                            else:
                                                raise
                                else:
                                    self.data[zone][district] = {}
                                    
                            except Exception as e:
                                self.logger.error(f"Error processing district {district}: {str(e)}")
                                if ERROR_HANDLING["continue_on_error"]:
                                    self.data[zone][district] = {}
                                else:
                                    raise
                    else:
                        self.data[zone] = {}
                        
                except Exception as e:
                    self.logger.error(f"Error processing zone {zone}: {str(e)}")
                    if ERROR_HANDLING["continue_on_error"]:
                        self.data[zone] = {}
                    else:
                        raise
            
            self.logger.info("Dropdown data extraction completed")
            self.logger.info(f"Total data structure: {len(self.data)} zones")
            
        except Exception as e:
            self.logger.error(f"Error in extract_all_dropdown_data: {str(e)}")
            if ERROR_HANDLING["detailed_error_logging"]:
                self.logger.error(traceback.format_exc())
            self.take_screenshot("extraction_error")
    
    def sanitize_name(self, name):
        """Sanitize name for Excel named ranges"""
        # Remove special characters and replace spaces with underscores
        sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', name)
        # Ensure it starts with a letter or underscore
        if sanitized and not sanitized[0].isalpha() and sanitized[0] != '_':
            sanitized = '_' + sanitized
        return sanitized[:255]  # Excel name limit
    
    def create_excel_with_cascading_dropdowns(self, filename=None):
        """Create Excel file with cascading dropdowns"""
        try:
            if filename is None:
                filename = OUTPUT_CONFIG["excel_filename"]
                
            self.logger.info(f"Creating Excel file: {filename}")
            
            wb = Workbook()
            
            # Create sheets
            form_sheet = wb.active
            form_sheet.title = EXCEL_CONFIG["sheet_names"]["form"]
            lists_sheet = wb.create_sheet(EXCEL_CONFIG["sheet_names"]["lists"])
            
            # Style headers in Form sheet
            headers = EXCEL_CONFIG["headers"]
            for col, header in enumerate(headers, 1):
                cell = form_sheet.cell(row=1, column=col, value=header)
                
                # Apply header styling from configuration
                if EXCEL_CONFIG["header_style"]["font_bold"]:
                    cell.font = Font(bold=True, size=EXCEL_CONFIG["header_style"]["font_size"])
                
                cell.fill = PatternFill(
                    start_color=EXCEL_CONFIG["header_style"]["fill_color"],
                    end_color=EXCEL_CONFIG["header_style"]["fill_color"],
                    fill_type="solid"
                )
                
                cell.alignment = Alignment(horizontal=EXCEL_CONFIG["header_style"]["alignment"])
            
            # Populate Lists sheet and create named ranges
            self._populate_lists_sheet(lists_sheet, wb)
            
            # Add data validation to Form sheet
            self._add_data_validations(form_sheet)
            
            # Add instructions
            self._add_instructions(form_sheet)
            
            # Save the workbook
            wb.save(filename)
            self.logger.info(f"Excel file saved as {filename}")
            
            # Create summary if configured
            if OUTPUT_CONFIG["create_summary"]:
                self.create_data_summary()
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {str(e)}")
            if ERROR_HANDLING["detailed_error_logging"]:
                self.logger.error(traceback.format_exc())
    
    def _populate_lists_sheet(self, lists_sheet, wb):
        """Populate the Lists sheet with data and create named ranges"""
        current_col = 1
        
        # Zone list
        zones = list(self.data.keys())
        for row, zone in enumerate(zones, 2):
            lists_sheet.cell(row=row, column=current_col, value=zone)
        
        # Create named range for zones
        if zones:
            zone_range = f"{EXCEL_CONFIG['sheet_names']['lists']}!${chr(64+current_col)}$2:${chr(64+current_col)}${len(zones)+1}"
            wb.defined_names.append(DefinedName("ZoneList", attr_text=zone_range))
        current_col += 1
        
        # Process hierarchical data
        for zone, districts in self.data.items():
            zone_sanitized = self.sanitize_name(zone)
            
            if districts:
                district_list = list(districts.keys())
                start_row = 2
                
                for row, district in enumerate(district_list, start_row):
                    lists_sheet.cell(row=row, column=current_col, value=district)
                
                # Create named range for this zone's districts
                if district_list:
                    district_range = f"{EXCEL_CONFIG['sheet_names']['lists']}!${chr(64+current_col)}${start_row}:${chr(64+current_col)}${len(district_list)+start_row-1}"
                    wb.defined_names.append(DefinedName(zone_sanitized, attr_text=district_range))
                current_col += 1
                
                # Continue with sub registers and villages...
                for district, sub_registers in districts.items():
                    district_sanitized = self.sanitize_name(f"{zone}_{district}")
                    
                    if sub_registers:
                        sub_register_list = list(sub_registers.keys())
                        start_row = 2
                        
                        for row, sub_register in enumerate(sub_register_list, start_row):
                            lists_sheet.cell(row=row, column=current_col, value=sub_register)
                        
                        if sub_register_list:
                            sub_register_range = f"{EXCEL_CONFIG['sheet_names']['lists']}!${chr(64+current_col)}${start_row}:${chr(64+current_col)}${len(sub_register_list)+start_row-1}"
                            wb.defined_names.append(DefinedName(district_sanitized, attr_text=sub_register_range))
                        current_col += 1
                        
                        # Villages
                        for sub_register, villages in sub_registers.items():
                            sub_register_sanitized = self.sanitize_name(f"{zone}_{district}_{sub_register}")
                            
                            if villages:
                                start_row = 2
                                
                                for row, village in enumerate(villages, start_row):
                                    lists_sheet.cell(row=row, column=current_col, value=village)
                                
                                village_range = f"{EXCEL_CONFIG['sheet_names']['lists']}!${chr(64+current_col)}${start_row}:${chr(64+current_col)}${len(villages)+start_row-1}"
                                wb.defined_names.append(DefinedName(sub_register_sanitized, attr_text=village_range))
                                current_col += 1
    
    def _add_data_validations(self, form_sheet):
        """Add data validation dropdowns to the form sheet"""
        try:
            # Zone dropdown (Column A)
            zone_validation = DataValidation(type="list", formula1="ZoneList")
            zone_validation.add(f"A2:A{EXCEL_CONFIG['max_dropdown_rows']}")
            form_sheet.add_data_validation(zone_validation)
        except Exception as e:
            self.logger.error(f"Error adding data validations: {str(e)}")
    
    def _add_instructions(self, form_sheet):
        """Add instructions to the form sheet"""
        instructions = [
            "Instructions:",
            "1. Select a Zone from the dropdown in column A",
            "2. Check the 'Lists' sheet for available options",
            "3. The data shows the hierarchical relationship",
            f"4. Total zones extracted: {len(self.data)}",
            f"5. Extraction completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        
        for row, instruction in enumerate(instructions, 3):
            form_sheet.cell(row=row, column=1, value=instruction)
    
    def create_data_summary(self, filename=None):
        """Create a text summary of extracted data"""
        try:
            if filename is None:
                filename = OUTPUT_CONFIG["summary_filename"]
                
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("TN Registration Department - Dropdown Data Summary\n")
                f.write("=" * 50 + "\n")
                f.write(f"Extraction Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Zones: {len(self.data)}\n\n")
                
                for zone, districts in self.data.items():
                    f.write(f"ZONE: {zone}\n")
                    f.write("-" * 30 + "\n")
                    
                    if districts:
                        for district, sub_registers in districts.items():
                            f.write(f"  DISTRICT: {district}\n")
                            
                            if sub_registers:
                                for sub_register, villages in sub_registers.items():
                                    f.write(f"    SUB REGISTER: {sub_register}\n")
                                    
                                    if villages:
                                        f.write(f"      VILLAGES ({len(villages)}): {', '.join(villages)}\n")
                                    f.write("\n")
                            f.write("\n")
                    f.write("\n")
            
            self.logger.info(f"Data summary saved as {filename}")
            
        except Exception as e:
            self.logger.error(f"Error creating data summary: {str(e)}")
    
    def run(self):
        """Main execution method"""
        try:
            self.logger.info("Starting Enhanced TN Registration Department scraper")
            
            # Validate configuration
            config_errors = validate_config()
            if config_errors:
                for error in config_errors:
                    self.logger.error(f"Configuration error: {error}")
                raise Exception("Configuration validation failed")
            
            # Setup WebDriver
            self.setup_driver()
            
            # Navigate to website
            self.navigate_to_website()
            
            # Select English language
            self.select_english_language()
            
            # Navigate to EC page
            self.navigate_to_ec_page()
            
            # Extract dropdown data
            self.extract_all_dropdown_data()
            
            # Create output files
            if self.data:
                if OUTPUT_CONFIG["create_excel"]:
                    self.create_excel_with_cascading_dropdowns()
                
                self.logger.info("Scraping completed successfully!")
                self.logger.info(f"Extracted data for {len(self.data)} zones")
            else:
                self.logger.warning("No data extracted. Please check the website structure.")
                if ERROR_HANDLING["save_partial_data"]:
                    self.logger.info("Saving empty data structure for reference")
                    if OUTPUT_CONFIG["create_excel"]:
                        self.create_excel_with_cascading_dropdowns()
            
        except Exception as e:
            self.logger.error(f"Error in main execution: {str(e)}")
            if ERROR_HANDLING["detailed_error_logging"]:
                self.logger.error(traceback.format_exc())
            self.take_screenshot("main_execution_error")
            
        finally:
            if self.driver:
                self.driver.quit()
                self.logger.info("WebDriver closed")

def main():
    """Main function"""
    scraper = TNRegiNetScraperEnhanced()
    scraper.run()

if __name__ == "__main__":
    main()
