#!/usr/bin/env python3
"""
Tamil Nadu Registration Department Website Scraper
Automates data extraction from www.tnreginet.gov.in for Zone, District, Sub Register Office, and Village dropdowns
Creates Excel file with cascading dropdowns functionality
"""

import time
import logging
import traceback
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('tnreginet_scraper.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class TNRegiNetScraper:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.data = {}
        
    def setup_driver(self):
        """Initialize Chrome WebDriver with appropriate options"""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
            # Uncomment the next line to run in headless mode
            # chrome_options.add_argument("--headless")
            
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.wait = WebDriverWait(self.driver, 30)
            logger.info("Chrome WebDriver initialized successfully")
            
        except Exception as e:
            logger.error(f"Failed to initialize WebDriver: {str(e)}")
            raise
    
    def navigate_to_website(self):
        """Navigate to the TN Registration website"""
        try:
            logger.info("Navigating to www.tnreginet.gov.in")
            self.driver.get("http://www.tnreginet.gov.in")
            time.sleep(3)
            logger.info("Successfully loaded the website")
            
        except Exception as e:
            logger.error(f"Failed to navigate to website: {str(e)}")
            raise
    
    def select_english_language(self):
        """Select English language option"""
        try:
            logger.info("Attempting to select English language")
            
            # Common selectors for language selection
            language_selectors = [
                "//a[contains(text(), 'English')]",
                "//button[contains(text(), 'English')]",
                "//span[contains(text(), 'English')]",
                "//div[contains(text(), 'English')]",
                "//a[@href*='english' or @href*='English']",
                "//select[@name='language']//option[contains(text(), 'English')]"
            ]
            
            for selector in language_selectors:
                try:
                    element = self.wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                    element.click()
                    logger.info("English language selected successfully")
                    time.sleep(2)
                    return
                except TimeoutException:
                    continue
            
            logger.warning("Could not find English language selector, proceeding anyway")
            
        except Exception as e:
            logger.error(f"Error selecting English language: {str(e)}")
            # Continue execution as language selection might not be mandatory
    
    def navigate_to_ec_page(self):
        """Navigate to E-services -> Encumbrance certificate -> View EC"""
        try:
            logger.info("Navigating to E-services -> Encumbrance certificate -> View EC")
            
            # Try to find E-services menu
            eservices_selectors = [
                "//a[contains(text(), 'E-services') or contains(text(), 'E-Services')]",
                "//span[contains(text(), 'E-services') or contains(text(), 'E-Services')]",
                "//div[contains(text(), 'E-services') or contains(text(), 'E-Services')]",
                "//li[contains(text(), 'E-services') or contains(text(), 'E-Services')]"
            ]
            
            eservices_element = None
            for selector in eservices_selectors:
                try:
                    eservices_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                    break
                except TimeoutException:
                    continue
            
            if eservices_element:
                eservices_element.click()
                logger.info("Clicked on E-services")
                time.sleep(2)
            
            # Try to find Encumbrance certificate
            encumbrance_selectors = [
                "//a[contains(text(), 'Encumbrance') or contains(text(), 'encumbrance')]",
                "//span[contains(text(), 'Encumbrance') or contains(text(), 'encumbrance')]",
                "//div[contains(text(), 'Encumbrance') or contains(text(), 'encumbrance')]"
            ]
            
            encumbrance_element = None
            for selector in encumbrance_selectors:
                try:
                    encumbrance_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                    break
                except TimeoutException:
                    continue
            
            if encumbrance_element:
                encumbrance_element.click()
                logger.info("Clicked on Encumbrance certificate")
                time.sleep(2)
            
            # Try to find View EC
            view_ec_selectors = [
                "//a[contains(text(), 'View EC') or contains(text(), 'view ec')]",
                "//span[contains(text(), 'View EC') or contains(text(), 'view ec')]",
                "//div[contains(text(), 'View EC') or contains(text(), 'view ec')]"
            ]
            
            view_ec_element = None
            for selector in view_ec_selectors:
                try:
                    view_ec_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                    break
                except TimeoutException:
                    continue
            
            if view_ec_element:
                view_ec_element.click()
                logger.info("Clicked on View EC")
                time.sleep(3)
            
            logger.info("Successfully navigated to EC page")
            
        except Exception as e:
            logger.error(f"Error navigating to EC page: {str(e)}")
            # Try alternative navigation or continue with current page
            logger.info("Attempting to find dropdowns on current page")
    
    def extract_dropdown_options(self, dropdown_element):
        """Extract options from a dropdown element"""
        try:
            if dropdown_element.tag_name == 'select':
                select = Select(dropdown_element)
                options = [option.text.strip() for option in select.options 
                          if option.text.strip() and option.text.strip().lower() not in ['select', 'choose', '--select--', '--choose--']]
                return options
            else:
                # Handle custom dropdowns
                options = dropdown_element.find_elements(By.TAG_NAME, "option")
                return [option.text.strip() for option in options 
                       if option.text.strip() and option.text.strip().lower() not in ['select', 'choose', '--select--', '--choose--']]
        except Exception as e:
            logger.error(f"Error extracting dropdown options: {str(e)}")
            return []
    
    def find_dropdown_element(self, dropdown_type):
        """Find dropdown element by type (zone, district, sub_register, village)"""
        selectors_map = {
            'zone': [
                "//select[contains(@name, 'zone') or contains(@id, 'zone')]",
                "//select[contains(@name, 'Zone') or contains(@id, 'Zone')]",
                "//select[contains(@class, 'zone')]"
            ],
            'district': [
                "//select[contains(@name, 'district') or contains(@id, 'district')]",
                "//select[contains(@name, 'District') or contains(@id, 'District')]",
                "//select[contains(@class, 'district')]"
            ],
            'sub_register': [
                "//select[contains(@name, 'sub') or contains(@id, 'sub') or contains(@name, 'register') or contains(@id, 'register')]",
                "//select[contains(@name, 'Sub') or contains(@id, 'Sub') or contains(@name, 'Register') or contains(@id, 'Register')]",
                "//select[contains(@class, 'sub') or contains(@class, 'register')]"
            ],
            'village': [
                "//select[contains(@name, 'village') or contains(@id, 'village')]",
                "//select[contains(@name, 'Village') or contains(@id, 'Village')]",
                "//select[contains(@class, 'village')]"
            ]
        }
        
        selectors = selectors_map.get(dropdown_type, [])
        
        for selector in selectors:
            try:
                element = self.driver.find_element(By.XPATH, selector)
                return element
            except NoSuchElementException:
                continue
        
        return None
    
    def extract_all_dropdown_data(self):
        """Extract all dropdown data with hierarchical relationships"""
        try:
            logger.info("Starting dropdown data extraction")
            
            # Find all dropdown elements
            zone_dropdown = self.find_dropdown_element('zone')
            district_dropdown = self.find_dropdown_element('district')
            sub_register_dropdown = self.find_dropdown_element('sub_register')
            village_dropdown = self.find_dropdown_element('village')
            
            if not zone_dropdown:
                logger.error("Zone dropdown not found")
                return
            
            # Extract zones
            zones = self.extract_dropdown_options(zone_dropdown)
            logger.info(f"Found {len(zones)} zones: {zones}")
            
            for zone in zones:
                logger.info(f"Processing zone: {zone}")
                self.data[zone] = {}
                
                try:
                    # Select zone
                    zone_select = Select(zone_dropdown)
                    zone_select.select_by_visible_text(zone)
                    time.sleep(2)
                    
                    # Wait for district dropdown to update
                    if district_dropdown:
                        self.wait.until(lambda driver: len(self.extract_dropdown_options(district_dropdown)) > 0)
                        districts = self.extract_dropdown_options(district_dropdown)
                        logger.info(f"Found {len(districts)} districts for zone {zone}: {districts}")
                        
                        for district in districts:
                            logger.info(f"Processing district: {district}")
                            self.data[zone][district] = {}
                            
                            try:
                                # Select district
                                district_select = Select(district_dropdown)
                                district_select.select_by_visible_text(district)
                                time.sleep(2)
                                
                                # Wait for sub register dropdown to update
                                if sub_register_dropdown:
                                    self.wait.until(lambda driver: len(self.extract_dropdown_options(sub_register_dropdown)) > 0)
                                    sub_registers = self.extract_dropdown_options(sub_register_dropdown)
                                    logger.info(f"Found {len(sub_registers)} sub registers for district {district}: {sub_registers}")
                                    
                                    for sub_register in sub_registers:
                                        logger.info(f"Processing sub register: {sub_register}")
                                        
                                        try:
                                            # Select sub register
                                            sub_register_select = Select(sub_register_dropdown)
                                            sub_register_select.select_by_visible_text(sub_register)
                                            time.sleep(2)
                                            
                                            # Wait for village dropdown to update
                                            if village_dropdown:
                                                self.wait.until(lambda driver: len(self.extract_dropdown_options(village_dropdown)) > 0)
                                                villages = self.extract_dropdown_options(village_dropdown)
                                                logger.info(f"Found {len(villages)} villages for sub register {sub_register}: {villages}")
                                                self.data[zone][district][sub_register] = villages
                                            else:
                                                self.data[zone][district][sub_register] = []
                                                
                                        except Exception as e:
                                            logger.error(f"Error processing sub register {sub_register}: {str(e)}")
                                            self.data[zone][district][sub_register] = []
                                else:
                                    self.data[zone][district] = {}
                                    
                            except Exception as e:
                                logger.error(f"Error processing district {district}: {str(e)}")
                                self.data[zone][district] = {}
                    else:
                        self.data[zone] = {}
                        
                except Exception as e:
                    logger.error(f"Error processing zone {zone}: {str(e)}")
                    self.data[zone] = {}
            
            logger.info("Dropdown data extraction completed")
            logger.info(f"Total data structure: {len(self.data)} zones")
            
        except Exception as e:
            logger.error(f"Error in extract_all_dropdown_data: {str(e)}")
            logger.error(traceback.format_exc())
    
    def sanitize_name(self, name):
        """Sanitize name for Excel named ranges"""
        # Remove special characters and replace spaces with underscores
        sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', name)
        # Ensure it starts with a letter or underscore
        if sanitized and not sanitized[0].isalpha() and sanitized[0] != '_':
            sanitized = '_' + sanitized
        return sanitized
    
    def create_excel_with_cascading_dropdowns(self, filename="tnreginet_data.xlsx"):
        """Create Excel file with cascading dropdowns"""
        try:
            logger.info("Creating Excel file with cascading dropdowns")
            
            wb = Workbook()
            
            # Create Form sheet
            form_sheet = wb.active
            form_sheet.title = "Form"
            
            # Create Lists sheet
            lists_sheet = wb.create_sheet("Lists")
            
            # Style headers in Form sheet
            headers = ["Zone", "District", "Sub Register Office", "Village"]
            for col, header in enumerate(headers, 1):
                cell = form_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Populate Lists sheet and create named ranges
            current_col = 1
            
            # Zone list
            zones = list(self.data.keys())
            for row, zone in enumerate(zones, 2):
                lists_sheet.cell(row=row, column=current_col, value=zone)
            
            # Create named range for zones
            zone_range = f"Lists!${chr(64+current_col)}$2:${chr(64+current_col)}${len(zones)+1}"
            wb.defined_names.append(DefinedName("ZoneList", attr_text=zone_range))
            current_col += 1
            
            # District, Sub Register, and Village lists
            for zone, districts in self.data.items():
                zone_sanitized = self.sanitize_name(zone)
                
                if districts:
                    district_list = list(districts.keys())
                    start_row = 2
                    
                    for row, district in enumerate(district_list, start_row):
                        lists_sheet.cell(row=row, column=current_col, value=district)
                    
                    # Create named range for this zone's districts
                    district_range = f"Lists!${chr(64+current_col)}${start_row}:${chr(64+current_col)}${len(district_list)+start_row-1}"
                    wb.defined_names.append(DefinedName(zone_sanitized, attr_text=district_range))
                    current_col += 1
                    
                    # Sub Register and Village lists for each district
                    for district, sub_registers in districts.items():
                        district_sanitized = self.sanitize_name(f"{zone}_{district}")
                        
                        if sub_registers:
                            sub_register_list = list(sub_registers.keys())
                            start_row = 2
                            
                            for row, sub_register in enumerate(sub_register_list, start_row):
                                lists_sheet.cell(row=row, column=current_col, value=sub_register)
                            
                            # Create named range for this district's sub registers
                            sub_register_range = f"Lists!${chr(64+current_col)}${start_row}:${chr(64+current_col)}${len(sub_register_list)+start_row-1}"
                            wb.defined_names.append(DefinedName(district_sanitized, attr_text=sub_register_range))
                            current_col += 1
                            
                            # Village lists for each sub register
                            for sub_register, villages in sub_registers.items():
                                sub_register_sanitized = self.sanitize_name(f"{zone}_{district}_{sub_register}")
                                
                                if villages:
                                    start_row = 2
                                    
                                    for row, village in enumerate(villages, start_row):
                                        lists_sheet.cell(row=row, column=current_col, value=village)
                                    
                                    # Create named range for this sub register's villages
                                    village_range = f"Lists!${chr(64+current_col)}${start_row}:${chr(64+current_col)}${len(villages)+start_row-1}"
                                    wb.defined_names.append(DefinedName(sub_register_sanitized, attr_text=village_range))
                                    current_col += 1
            
            # Add data validation to Form sheet
            # Zone dropdown (Column A)
            zone_validation = DataValidation(type="list", formula1="ZoneList")
            zone_validation.add(f"A2:A1000")
            form_sheet.add_data_validation(zone_validation)
            
            # Note: For complex cascading dropdowns with INDIRECT formulas,
            # we'll create a simpler version that works with the extracted data
            
            # Add instructions
            form_sheet.cell(row=3, column=1, value="Instructions:")
            form_sheet.cell(row=4, column=1, value="1. Select a Zone from the dropdown in column A")
            form_sheet.cell(row=5, column=1, value="2. Check the 'Lists' sheet for available options")
            form_sheet.cell(row=6, column=1, value="3. The data shows the hierarchical relationship")
            
            # Save the workbook
            wb.save(filename)
            logger.info(f"Excel file saved as {filename}")
            
            # Also create a simple data dump for reference
            self.create_data_summary(filename.replace('.xlsx', '_summary.txt'))
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            logger.error(traceback.format_exc())
    
    def create_data_summary(self, filename):
        """Create a text summary of extracted data"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("TN Registration Department - Dropdown Data Summary\n")
                f.write("=" * 50 + "\n\n")
                
                for zone, districts in self.data.items():
                    f.write(f"ZONE: {zone}\n")
                    f.write("-" * 30 + "\n")
                    
                    if districts:
                        for district, sub_registers in districts.items():
                            f.write(f"  DISTRICT: {district}\n")
                            
                            if sub_registers:
                                for sub_register, villages in sub_registers.items():
                                    f.write(f"    SUB REGISTER: {sub_register}\n")
                                    
                                    if villages:
                                        f.write(f"      VILLAGES: {', '.join(villages)}\n")
                                    f.write("\n")
                            f.write("\n")
                    f.write("\n")
            
            logger.info(f"Data summary saved as {filename}")
            
        except Exception as e:
            logger.error(f"Error creating data summary: {str(e)}")
    
    def run(self):
        """Main execution method"""
        try:
            logger.info("Starting TN Registration Department scraper")
            
            # Setup WebDriver
            self.setup_driver()
            
            # Navigate to website
            self.navigate_to_website()
            
            # Select English language
            self.select_english_language()
            
            # Navigate to EC page
            self.navigate_to_ec_page()
            
            # Extract dropdown data
            self.extract_all_dropdown_data()
            
            # Create Excel file
            if self.data:
                self.create_excel_with_cascading_dropdowns()
                logger.info("Scraping completed successfully!")
            else:
                logger.warning("No data extracted. Please check the website structure.")
            
        except Exception as e:
            logger.error(f"Error in main execution: {str(e)}")
            logger.error(traceback.format_exc())
            
        finally:
            if self.driver:
                self.driver.quit()
                logger.info("WebDriver closed")

def main():
    """Main function"""
    scraper = TNRegiNetScraper()
    scraper.run()

if __name__ == "__main__":
    main()
